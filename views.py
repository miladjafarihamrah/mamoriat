from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate
from django.conf import settings
from django.http import HttpResponse
from django.contrib import messages
from django.shortcuts import get_object_or_404

from .forms import SignUpForm, LoginForm
from .forms import MissionForm
from .forms import ExpenseForm
from .forms import ReportForm

from .models import Mission
from .models import Expense  
from .models import Balance


import jdatetime
from django.http import JsonResponse
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook


# دیکشنری برای تبدیل نام ماه‌های شمسی به انگلیسی
PERSIAN_MONTH_TO_ENGLISH = {
    'فروردین': 'Farvardin',
    'اردیبهشت': 'Ordibehesht',
    'خرداد': 'Khordad',
    'تیر': 'Tir',
    'مرداد': 'Mordad',
    'شهریور': 'Shahrivar',
    'مهر': 'Mehr',
    'آبان': 'Aban',
    'آذر': 'Azar',
    'دی': 'Dey',
    'بهمن': 'Bahman',
    'اسفند': 'Esfand'
}


def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('home')
    else:
        form = SignUpForm()
    return render(request, 'signup.html', {'form': form})

def user_login(request):
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
    else:
        form = LoginForm()
    return render(request, 'login.html', {'form': form})

def home(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    # دریافت تاریخ جاری به صورت شمسی
    today = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    # استخراج ماه جاری (به صورت عددی)
    current_month = jdatetime.datetime.now().month  # ماه جاری به صورت عدد (1 تا 12)
    
    # فیلتر کردن ماموریت‌ها بر اساس ماه جاری
    missions = Mission.objects.filter(user=request.user)
    missions_this_month = [mission for mission in missions if int(mission.date.split('/')[1]) == current_month]
    
    # شمارش تعداد ماموریت‌های ماه جاری
    mission_count = len(missions_this_month)
    full_name = f"{request.user.first_name} {request.user.last_name}"
                
    # دریافت مانده حساب
    try:
        balance = Balance.objects.get(user=request.user).amount
    except Balance.DoesNotExist:
        # اگر مانده حساب وجود نداشت، یک مانده حساب جدید ایجاد کنید
        balance = 4000000  # مقدار پیش‌فرض
        Balance.objects.create(user=request.user, amount=balance)

    return render(request, 'home.html', {
        'full_name': full_name,
        'today': today,
        'mission_count': mission_count,  # تعداد ماموریت‌های ماه جاری
        'balance': balance,  # ارسال مانده حساب به تمپلیت
    })

import os
from openpyxl import Workbook, load_workbook

def add_mission(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    if request.method == 'POST':
        form = MissionForm(request.POST)
        if form.is_valid():
            mission = form.save(commit=False)
            mission.user = request.user

            # بررسی تکراری نبودن تاریخ ماموریت
            existing_mission = Mission.objects.filter(user=request.user, date=mission.date).exists()
            if existing_mission:
                messages.error(request, 'خطا: تاریخ ماموریت تکراری است.')
                return render(request, 'add_mission.html', {'form': form})

            mission.save()  # ذخیره در دیتابیس
            return redirect('home')
        else:
            messages.error(request, 'خطا در ثبت ماموریت. لطفاً دوباره تلاش کنید.')
    else:
        form = MissionForm()
    return render(request, 'add_mission.html', {'form': form})
def edit_mission(request):
    month = request.GET.get('month')  # دریافت ماه از پارامتر URL
    if not month:
        return render(request, 'error.html', {'message': 'ماه معتبر نیست.'})

    # فیلتر کردن ماموریت‌ها بر اساس ماه (برای فیلد CharField)
    missions = Mission.objects.filter(user=request.user)
    missions = [mission for mission in missions if mission.date.split('/')[1] == month]

    return render(request, 'edit_mission.html', {'missions': missions, 'month': month})

def delete_mission(request):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'احراز هویت لازم است.'})
    
    date = request.GET.get('date')  # دریافت تاریخ از درخواست GET
    if not date:
        return JsonResponse({'status': 'error', 'message': 'تاریخ نامعتبر است.'})

    # حذف ماموریت از دیتابیس
    mission = get_object_or_404(Mission, date=date, user=request.user)
    mission.delete()

    return JsonResponse({'status': 'success', 'message': 'ماموریت با موفقیت حذف شد.'})


  


def add_expense(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    if request.method == 'POST':
        form = ExpenseForm(request.POST)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.user = request.user
            expense.save()

            # کاهش مانده حساب
            balance = Balance.objects.get(user=request.user)
            balance.amount -= expense.amount
            balance.save()

        
            return redirect('home')
        else:
            messages.error(request, 'خطا در ثبت تنخواه. لطفاً دوباره تلاش کنید.')
    else:
        form = ExpenseForm()
    return render(request, 'add_expense.html', {'form': form})

# ویرایش هزینه‌ها
def edit_expense(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    month = request.GET.get('month')  # دریافت ماه از پارامتر URL
    if not month:
        return render(request, 'error.html', {'message': 'ماه معتبر نیست.'})

    # فیلتر کردن هزینه‌ها بر اساس ماه
    expenses = Expense.objects.filter(user=request.user)
    expenses = [expense for expense in expenses if expense.date.split('/')[1] == month]

    return render(request, 'edit_expense.html', {'expenses': expenses, 'month': month})



def delete_expense(request):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'احراز هویت لازم است.'})
    
    expense_id = request.GET.get('id')  # دریافت id به جای date

    if not expense_id:
        return JsonResponse({'status': 'error', 'message': 'شناسه نامعتبر است.'})

    # حذف Expense بر اساس ID
    expense = get_object_or_404(Expense, id=expense_id, user=request.user)
    amount = expense.amount  # مقدار هزینه برای به‌روزرسانی مانده حساب
    expense.delete()

    # افزایش مانده حساب
    balance = Balance.objects.get(user=request.user)
    balance.amount += amount
    balance.save()

    return JsonResponse({'status': 'success', 'message': 'تنخواه با موفقیت حذف شد.'})

from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from datetime import datetime
from .forms import ReportForm
from .models import Mission, Expense

def generate_report(request):
    # بررسی احراز هویت کاربر
    if not request.user.is_authenticated:
        return redirect('login')
    
    if request.method == 'POST':
        form = ReportForm(request.POST)
        if form.is_valid():
            try:
                # دریافت و اعتبارسنجی ماه
                month = int(form.cleaned_data['month'])
                
                if not (1 <= month <= 12):
                    messages.error(request, 'ماه وارد شده نامعتبر است.')
                    return redirect('home')
            except (ValueError, KeyError):
                messages.error(request, 'خطا در پردازش داده‌ها.')
                return redirect('home')

            # دریافت نوع گزارش
            report_type = request.POST.get('report_type')
            REPORT_TYPES = ('mission', 'expense')

            # بررسی معتبر بودن نوع گزارش
            if report_type not in REPORT_TYPES:
                messages.error(request, 'نوع گزارش نامعتبر است.')
                return redirect('home')

            # تنظیمات مربوط به نوع گزارش
            report_config = {
                'mission': {
                    'model': Mission,
                    'output_name': f"m_{request.user.username}_{month}.xlsx",
                    'header': ['ردیف', 'تاریخ', 'کارخانه']
                },
                'expense': {
                    'model': Expense,
                    'output_name': f"t_{request.user.username}_{month}.xlsx",
                    'header': ['ردیف', 'تاریخ', 'توضیحات', 'مبلغ', 'کارخانه']
                }
            }

            config = report_config[report_type]
            model = config['model']
            output_name = config['output_name']
            header = config['header']

            # دریافت داده‌ها از دیتابیس
            data = model.objects.filter(user=request.user)
            filtered_data = []
            total_amount = 0  # جمع کل مبالغ

            for item in data:
                item_date = item.date  # تاریخ به صورت رشته (مثلاً '1402/07/15')
                item_month = int(item_date.split('/')[1])  # استخراج ماه از تاریخ
                
                # فیلتر کردن داده‌ها بر اساس ماه
                if item_month == month:
                    # تبدیل تاریخ به فرمت datetime برای مرتب‌سازی
                    date_obj = datetime.strptime(item_date, '%Y/%m/%d')
                    
                    if report_type == 'mission':
                        filtered_data.append([date_obj, item.date, item.factory])
                    elif report_type == 'expense':
                        # محاسبه جمع کل مبالغ
                        total_amount += item.amount
                        filtered_data.append([date_obj, item.date, item.description, item.amount, item.factory])

            # مرتب‌سازی داده‌ها بر اساس تاریخ (صعودی)
            filtered_data.sort(key=lambda x: x[0])  # x[0] همان date_obj است

            # حذف date_obj از داده‌ها (چون فقط برای مرتب‌سازی استفاده شد)
            filtered_data = [item[1:] for item in filtered_data]

            # ایجاد فایل اکسل جدید
            wb = Workbook()
            ws = wb.active
            ws.sheet_view.rightToLeft = True  # تنظیم جهت صفحه به راست‌چین
            ws.append(header)

            # تنظیم استایل برای هدرها
            header_font = Font(bold=True, color="FFFFFF")  # فونت پررنگ و رنگ سفید
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # رنگ پس‌زمینه آبی
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )  # حاشیه نازک برای سلول‌ها
            alignment = Alignment(horizontal='center', vertical='center')  # تراز وسط

            # اعمال استایل به هدرها
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = alignment

            # اضافه کردن داده‌ها به فایل اکسل همراه با شماره ردیف
            for index, row in enumerate(filtered_data, start=1):
                ws.append([index] + row)  # اضافه کردن شماره ردیف به ابتدای سطر

            # اعمال استایل به داده‌ها
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(header)):
                for cell in row:
                    cell.border = border
                    cell.alignment = alignment

            # اضافه کردن سطر جمع کل (فقط برای گزارش‌های expense و اگر داده وجود دارد)
            if report_type == 'expense' and filtered_data:
                ws.append(['', '', 'جمع کل', total_amount, ''])
                # اعمال استایل به سطر جمع کل
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True)
                    cell.border = border
                    cell.alignment = alignment
            full_name = f"{request.user.first_name} {request.user.last_name}"
        
            # اضافه کردن جای امضا در پایین صفحه
            ws.append([])  # یک سطر خالی
            ws.append(['','مدیر عامل',full_name, 'کنترل کننده', ])  # جای امضای کاربر
           
            # تنظیم خودکار اندازه ستون‌ها
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # نام ستون (مثلاً A, B, C)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2  # محاسبه عرض ستون
                ws.column_dimensions[column].width = adjusted_width

            # تنظیمات صفحه برای پرینت
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # جهت صفحه به صورت افقی
            ws.page_setup.paperSize = ws.PAPERSIZE_A4  # اندازه کاغذ A4
            ws.page_setup.fitToWidth = 1  # تنظیم عرض صفحه برای پرینت
            ws.page_setup.fitToHeight = 0  # عدم تنظیم ارتفاع صفحه

            # تنظیم حاشیه‌ها
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.top = 0.5
            ws.page_margins.bottom = 0.5
            ws.page_margins.header = 0.3
            ws.page_margins.footer = 0.3

            # تنظیم هدر و فوتر
            ws.oddHeader.center.text = "گزارش ماهانه"  # هدر وسط صفحه
            ws.oddFooter.center.text = "صفحه &P از &N"  # فوتر وسط صفحه (شماره صفحه)

            # تنظیم هدرهای HTTP برای دانلود فایل
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{output_name}"'

            # ذخیره فایل در پاسخ
            wb.save(response)
            return response
    else:
        form = ReportForm()
    return render(request, 'report.html', {'form': form})

from django.contrib import messages
from .models import Balance

def update_balance(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    if request.method == 'POST':
        amount = int(request.POST.get('amount', 0))
        action = request.POST.get('action')  # 'increase' یا 'decrease'

        if amount <= 0:
            messages.error(request, 'مبلغ وارد شده نامعتبر است.')
            return redirect('home')

        balance = Balance.objects.get(user=request.user)
        if action == 'increase':
            balance.amount += amount
        elif action == 'decrease':
            balance.amount -= amount
        balance.save()

        messages.success(request, f'مانده حساب با موفقیت {"افزایش" if action == "increase" else "کاهش"} یافت.')
        return redirect('home')
    else:
        return redirect('home')




